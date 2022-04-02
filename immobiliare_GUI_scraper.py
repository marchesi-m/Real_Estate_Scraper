# -*- coding: utf-8 -*-
"""
@author: Martina
"""

import tkinter as tk
import bs4
import requests
import openpyxl
from openpyxl import Workbook
import pathlib
import webbrowser
from tkinter import ttk
import json
from json import JSONDecodeError
import os


# SETTING TEXT VARIABLES
aaa = 'Arial Bold'
warn = "Warning! You have not chosen the file type."

warn1 = """Warning! The JSON file generation was not successful
due to an "Object of type 'type' is not JSON serializable" error.
The dictionary output has one item which cannot be serialized 
in a JSON file."""

warn2 = """Warning! The current search results could not be added
 due to an "JSONDecodeError". The search results have been
 converted to JSON correctly, however they could not be appended
 due to the current JSON file's incorrect format."""

instructions1 = """1.  Click the link to the Immobiliare.it website
2.  Select a LOCATION in the search bar and click 'SEARCH'
3.  A webpage will appear similar to the one below"""

instructions2 = """4.  Select whether you want to
     BUY or RENT a property
5.  Change the 'MOST RELEVANT' filter
     (if needed)
6.  Click the 'SEARCH' button
7.  Paste the link in the 'URL' entry box
8.  Select which file NAME and TYPE you
     would like to save
9.  Click 'SCRAPE IT!' and the files will
     be placed in the same directory as the
     python file."""

instructions4 = """⚠ IF THE FILE NAME IS NOT IN THE CURRENT DIRECTORY,
 A NEW FILE WILL BE CREATED. IF THE FILE ALREADY EXISTS
 THE NEW CONTENT WILL BE ADDED TO IT. ⚠"""
 
resu = """Your current search didn't yield any results.
Try changing your parameters."""


# CREATING THE MAIN WINDOW
window = tk.Tk()
window.title("Real Estate Scraping Tool")
window.geometry("439x833")
window.resizable(width=0, height=0)


# FUNCTIONS
def callback(webpage):
    """
    This function allows the user to interact with the label by clicking
    its link.
    """
    webbrowser.open_new(webpage)


def how_to():
    """
    This function creates another window with the program's instructions.
    """
    instruction_win = tk.Toplevel()
    instruction_win.title("Instructions")
    instruction_win.minsize(950, 770)
    instruction_win.grid_rowconfigure(0, weight=1)
    instruction_win.grid_rowconfigure(1, weight=1)
    instruction_win.grid_rowconfigure(2, weight=1)
    instruction_win.grid_rowconfigure(3, weight=1)
    instruction_win.grid_rowconfigure(4, weight=1)
    instruction_win.grid_columnconfigure(0, weight=1)
    instruction_win.grid_columnconfigure(1, weight=1)

    # instructions for use 1
    instructions_lbl1 = tk.Label(instruction_win, text=instructions1,
                                 font=(aaa, 13), justify='left')
    instructions_lbl1.grid(row=0, column=0, columnspan=2, padx=10, ipadx=20,
                           pady=(20, 10), sticky='W')

    # picture 1
    photo1 = tk.PhotoImage(file="guide_img1.png")
    photo_lbl1 = tk.Label(instruction_win, image=photo1, borderwidth=5,
                          relief='ridge')
    photo_lbl1.grid(row=1, column=0, columnspan=3, padx=20, pady=10)

    # instructions for use 2
    instructions_lbl2 = tk.Label(instruction_win, text=instructions2,
                                 font=(aaa, 13), justify='left')
    instructions_lbl2.grid(row=2, column=0, columnspan=1, padx=10, pady=10)

    # picture 2
    photo2 = tk.PhotoImage(file="guide_img2.png")
    photo_lbl2 = tk.Label(instruction_win, image=photo2, borderwidth=5,
                          relief='ridge')
    photo_lbl2.grid(row=2, column=1, columnspan=2, padx=(10, 20), pady=10)

    # instructions for use 4
    instructions_lbl4 = tk.Label(instruction_win, text=instructions4,
                                 font=(aaa, 13), justify='center',
                                 borderwidth=5, relief='ridge', bg='#fcdb03')
    instructions_lbl4.grid(row=3, column=0, columnspan=3, padx=10, pady=30,
                           ipadx=20, ipady=10)

    instruction_win.mainloop()


def search_page():
    """
    This function takes in a link from a SEARCH in immobiliare.it and collects
    all data from the listings in the page using the beautiful soup module.
    It returns a dictionary of dictionaries.
    """

    # SETTING UP THE PROGRAM WITH THE URL AND SOUP OBJECT
    url = URL.get()
    my_url = requests.get(url)
    soup = bs4.BeautifulSoup(my_url.text, "lxml")

    # GET INFORMATION PACKET, THEN STRIP IT AND ADD TO DICT INSIDE BIG LOOP
    packets = soup.find_all('li',
                            class_='nd-list__item in-realEstateResults__item')

    # SETTING THE CATEGORY (are the properties for sale, rent or auction?)
    buy = "vendita-case"
    rent = "affitto-case"
    auction = "aste-immobiliari"

    prop_category = ""

    if buy in url:
        prop_category = "BUY"
    elif rent in url:
        prop_category = "RENT"
    elif auction in url:
        prop_category = "AUCTION"
    else:
        return 'No category, INVALID URL'

    # CREATE THE EMPTY DICTIONARY OF DICTIONARIES (VERY IMPORTANT)
    data_dict = {}

    # LOOPING OVER EACH LISTING
    for packet in packets:

        # SETTING EMPTY VARIABLES FOR EACH PACKET
        packet_title = str          # title of listing
        packet_pricevar = int       # max price value
        packet_bathvar = int        # min bath var
        packet_surfvar = int        # min surface var
        packet_roomvar = int        # min rooms var
        packet_link = str           # link to listing
        packet_category = str       # category of listing

        # GET LISTING TITLES
        packet_title = packet.find('a', class_='in-card__title').text.strip()

        # GET CATEGORY category = PROP_CATEGORY --> add to dict
        packet_category = prop_category

        # GET LISTING LINK
        packet_link = packet.a.get('href')

        # GET PRICE
        price = packet.find('li', class_='nd-list__item in-feat__item in-feat__item--main in-realEstateListCard__features--main').text.strip()

        # take care of slash in case of rent and remove euro symbol
        clear_price_str = price[price.index('€')+1:].strip()

        if '/' in clear_price_str:
            clear_price_str = clear_price_str[:clear_price_str.index('/')].strip()

        packet_pricevar = clear_price_str.replace(',', '')  # can do without if

        if '€' in packet_pricevar:
            packet_pricevar = int(packet_pricevar[:packet_pricevar.index('€')].strip())
        else:
            packet_pricevar = int(packet_pricevar)

        # get items in li tags inside packet
        listing_features = packet.find_all('li', class_='nd-list__item in-feat__item')

        # iterating over tags that contain the house features
        for num, feature in enumerate(listing_features):

            # strings are converted into integers with some customization
            feat_string = feature.text

            if num == 0:
                mod_string = feat_string[:1].strip()
                packet_roomvar = int(mod_string)
                # print(f"ROOMS = {feat_string}")

            if num == 1:
                mod_string = feat_string[:feat_string.index('m')].strip()
                packet_surfvar = int(mod_string)
                # print(f"SURFACE = {packet_surfvar}")

            if num == 2:
                mod_string = feat_string[:1].strip()
                packet_bathvar = int(mod_string)
                # print(f"BATHROOMS = {packet_bathvar}")

        # AFTER COLLECTING THE OBJECTS, DECIDE BASED ON PARAMETERS IF THEY FIT
        rentvar_p = int(rentvar.get())      # max rent value from GUI
        buyvar_p = int(buyvar.get())        # max buy value from GUI
        bathvar_p = int(bathvar.get())      # min bath var from GUI
        surfvar_p = int(surfvar.get())      # min surface vat from GUI
        roomvar_p = int(roomvar.get())      # min rooms var from GUI

        # ADD A STRIKE FOR EACH SELECTED CONDITION THAT IS NOT MET
        strikes = 0

        if packet_category == "BUY" or packet_category == "AUCTION":
            if buyvar_p < packet_pricevar:
                strikes += 1
        if packet_category == "RENT":
            if rentvar_p < packet_pricevar:
                strikes += 1
        if packet_bathvar < bathvar_p:
            strikes += 1
        if packet_surfvar < surfvar_p:
            strikes += 1
        if packet_roomvar < roomvar_p:
            strikes += 1

        if strikes == 0:
            packet_dictionary = {
                    'listing_title' : packet_title,
                    'category' : packet_category,
                    'listing_link': packet_link,
                    'price' : packet_pricevar,
                    'surface' : packet_surfvar,
                    'rooms' : packet_roomvar,
                    'bathrooms' : packet_bathvar
                    }

            # ADD ALL THE RELEVANT DATA TO THE DICTIONARY
            data_dict[packet_title] = packet_dictionary

    print(f"Dictionary: {data_dict}\n")        # to check dictionary output
    return data_dict


def add_to_excel(dict_output):
    """
    This function adds the characteristics given by the other functions and
    adds them to a selected Excel file. There are three sheets in one workbook.
    """

    # excel file name generated by entry
    filename = filename_entry.get()
    file_excel = f"{filename}"+".xlsx"

    file_use = pathlib.Path(file_excel)

    if not os.path.exists(file_use):
        # creating excel file and assigning column headers
        file = Workbook()
        sheet = file.active

        sheet["A1"] = "LISTING TITLE"
        sheet["B1"] = "CATEGORY"
        sheet["C1"] = "PRICE (€)"
        sheet["D1"] = "ROOMS"
        sheet["E1"] = "SURFACE (m^2)"
        sheet["F1"] = "BATHROOMS"
        sheet["G1"] = "PAGE LINK"
        sheet["H1"] = "END ROW"

        file.save(file_excel)

    # loading and opening the file
    file = openpyxl.load_workbook(file_excel)
    sheet = file.active

    for k, val in dict_output.items():
        # add invisible row at the end (needed by the program to set max row)
        sheet.cell(column=8, row=sheet.max_row+1, value="end")

        for key, value in val.items():
            try:
                if key == 'listing_title':
                    sheet.cell(column=1, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'listing_title':
                    sheet.cell(column=1, row=sheet.max_row, value="n/a")
            try:
                if key == 'category':
                    sheet.cell(column=2, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'category':
                    sheet.cell(column=2, row=sheet.max_row, value="n/a")
            try:    
                if key == 'price':
                    sheet.cell(column=3, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'price':
                    sheet.cell(column=3, row=sheet.max_row, value="n/a")
            try:
                if key == 'rooms':
                    sheet.cell(column=4, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'rooms':
                    sheet.cell(column=4, row=sheet.max_row, value="n/a")
            try:
                if key == 'surface':
                    sheet.cell(column=5, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'surface':
                    sheet.cell(column=5, row=sheet.max_row, value="n/a")
            try:
                if key == 'bathrooms':
                    sheet.cell(column=6, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'bathrooms':
                    sheet.cell(column=6, row=sheet.max_row, value="n/a")
            try:
                if key == 'listing_link':
                    sheet.cell(column=7, row=sheet.max_row, value=value)
            except ValueError:
                if key == 'listing_link':
                    sheet.cell(column=7, row=sheet.max_row, value="n/a")
    
    file.save(file_excel)


def add_to_json(dict_output):
    """
    This function takes the dictionary output of the search_page function and
    inserts it in a JSON file. If the file does not exist, it will be created.
    The format of the dictionary is a list containing dictionaries, and each
    dictionary represents the results of a search.
    """

    # json file name generated by file name entry box
    filename = filename_entry.get()
    file = f"{filename}"+".json"

    # if path does not exist, create it
    if not os.path.exists(file):

        with open(file, 'w') as f:
            json.dump([], f)

    # if file is raising a type erorr then it has to be deleted
    with open(file, "r+") as f:
        try:
            data = json.load(f)
            data.append(dict_output)
            f.seek(0)
            json.dump(data, f)
        except TypeError:
            tk.messagebox.showerror(title="Error", message=warn1)
        except JSONDecodeError:
            tk.messagebox.showerror(title="Error", message=warn2)


def add_to_text(dict_output):
    """
    This function adds the dictionary given by search_page and writes it to
    a text file with formatting to make the information more readable.
    """
    # text file name generated by entry
    filename = filename_entry.get()
    file = f"{filename}"+".txt"

    txt_output = ""

    for k, val in dict_output.items():

        txt_output += "\n"

        for key, value in val.items():   # dict_output[dic]:
            if key == 'listing_title':
                txt_output += 'LISTING TITLE: ' + str(value) + "\n"
            elif key == 'category':
                txt_output += 'CATEGORY: ' + str(value) + "\n"
            elif key == 'listing_link':
                txt_output += 'LINK: ' + str(value) + "\n"
            elif key == 'price':
                txt_output += 'PRICE: ' + str(value) + "\n"
            elif key == 'surface':
                txt_output += 'SURFACE: ' + str(value) + "\n"
            elif key == 'bathrooms':
                txt_output += 'BATHROOMS: ' + str(value) + "\n"
            elif key == 'rooms':
                txt_output += 'ROOMS: ' + str(value) + "\n"
        txt_output += "\n"

    # append or write to file the dictionary content
    if os.path.exists(file):
        with open(file, 'a') as f:      # encoding="utf-8" if needed
            f.write(txt_output)
    else:
        with open(file, 'w') as f:
            f.write(txt_output)


def scrape_master():
    """
    Uses all the previous functions to produce the final result: appending
    the selected features to the file type chosen. This function is
    associated to the scraping button.
    """
    # search the page and make the dictionary output
    results = search_page()

    # checking the output dictionary is not empty
    if results != {}:
        if txtvar.get() == 1:
            add_to_text(results)        # add to text file

        if excelvar.get() == 1:
            add_to_excel(results)       # add to excel file

        if jsonvar.get() == 1:
            add_to_json(results)        # add to json file

        if (txtvar.get() == 0 and jsonvar.get() == 0 and excelvar.get() == 0):
            return tk.messagebox.showerror(title="Error", message=warn)
    else:
        return tk.messagebox.showinfo(title="No Results", message=resu)


# MAKING THE BASIC FEATURES --------------------------------------------------

# link to website
immobiliare_link_lbl = tk.Label(window, text="Immobiliare.it  🏠",
                                cursor="hand2", borderwidth=3, font=(aaa, 20),
                                relief="sunken", bg="#3c6a85", fg="white")
immobiliare_link_lbl.grid(row=0, column=0, columnspan=3, padx=10, pady=20,
                          ipady=10, ipadx=15)
immobiliare_link_lbl.bind("<Button-1>",
                          lambda x: callback("https://www.immobiliare.it/en/"))

# how to use this program button
howto_btn = tk.Button(window, text="How to use this program?", command=how_to,
                      width=30, font=(aaa, 12), bg='#cccccc')
howto_btn.grid(row=1, column=0, columnspan=3)

# url input and its label
URL_lbl = tk.Label(window, text="URL", font=(aaa, 15))
URL_lbl.grid(row=2, column=0, padx=(20, 0), pady=30, sticky='WE')

URL = tk.StringVar(window)  # , value="Paste URL here")
link_entry = tk.Entry(window, textvariable=URL, width=33, font=(aaa, 13))
link_entry.grid(row=2, column=1, columnspan=2, padx=(20, 40), pady=35, ipady=5,
                ipadx=5)

# frame additional filters
filters_frame = tk.LabelFrame(window, text="Additional Filters",
                              font=(aaa, 12))
filters_frame.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

# price, bathrooms, surface, rooms + selectboxes
rent_price_lbl = tk.Label(filters_frame, text="Max renting budget",
                          font=(aaa, 13))
rent_price_lbl.grid(row=0, column=0, padx=10, pady=10, sticky='E')
rent_price_lbl1 = tk.Label(filters_frame, text="€/month", font=(aaa, 13))
rent_price_lbl1.grid(row=0, column=2, padx=10, pady=10, sticky='W')
rentvar = tk.StringVar()
rentvar.set("1000")
rent_spin = ttk.Spinbox(filters_frame, from_=100.0, to=10_000.0,
                        increment=100.0, textvariable=rentvar, font=(aaa, 13),
                        width=10)
rent_spin.grid(row=0, column=1, padx=10, pady=10)

buy_price_lbl = tk.Label(filters_frame, text="Max buying budget",
                         font=(aaa, 13))
buy_price_lbl.grid(row=1, column=0, padx=10, pady=10, sticky='E')
buy_price_lbl2 = tk.Label(filters_frame, text="€", font=(aaa, 13))
buy_price_lbl2.grid(row=1, column=2, padx=10, pady=10, sticky='W')
buyvar = tk.StringVar()
buyvar.set("375000")
buy_spin = ttk.Spinbox(filters_frame, from_=25_000.0, to=1_500_0000.0,
                       increment=25_000.0, textvariable=buyvar,
                       font=(aaa, 13), width=10)
buy_spin.grid(row=1, column=1, padx=10, pady=10)

bathrooms_lbl = tk.Label(filters_frame,
                         text="Min. bathrooms",
                         font=(aaa, 13))
bathrooms_lbl.grid(row=3, column=0, padx=10, pady=10, sticky='E')

bathvar = tk.StringVar()
bathvar.set("1")
buy_spin = ttk.Spinbox(filters_frame, from_=1, to=4, increment=1,
                       textvariable=bathvar, font=(aaa, 13), width=10)
buy_spin.grid(row=3, column=1, padx=10, pady=10)

surface_lbl = tk.Label(filters_frame,
                       text="Min. surface",
                       font=(aaa, 13))
surface_lbl.grid(row=2, column=0, padx=10, pady=10, sticky='E')
surface_lbl1 = tk.Label(filters_frame, text="m^2", font=(aaa, 13))
surface_lbl1.grid(row=2, column=2, padx=10, pady=10, sticky='W')

surfvar = tk.StringVar()
surfvar.set("60")
surf_spin = ttk.Spinbox(filters_frame, from_=30, to=350, increment=10,
                        textvariable=surfvar, font=(aaa, 13), width=10)
surf_spin.grid(row=2, column=1, padx=10, pady=10)

rooms_lbl = tk.Label(filters_frame, text="Min. rooms", font=(aaa, 13))
rooms_lbl.grid(row=4, column=0, padx=10, pady=10, sticky='E')
roomvar = tk.StringVar()
roomvar.set("2")
room_spin = ttk.Spinbox(filters_frame, from_=1, to=8, increment=1,
                        textvariable=roomvar, font=(aaa, 13), width=10)
room_spin.grid(row=4, column=1, padx=10, pady=10)

# frame save options
save_frame = tk.LabelFrame(window, text="Save Options", font=(aaa, 12))
save_frame.grid(row=4, column=0, columnspan=3, padx=20, pady=10)

filename_lbl = tk.Label(save_frame, text="Filename", font=(aaa, 13))
filename_lbl.grid(row=0, column=0, padx=10, pady=10, sticky='E')

default_file = tk.StringVar(save_frame, value="House_Search")
filename_entry = tk.Entry(save_frame, textvariable=default_file, width=30,
                          font=(aaa, 13))
filename_entry.grid(row=0, column=1, columnspan=1, padx=10, pady=10, ipady=5,
                    ipadx=5)

# excel, JSON and text label + check boxes
excelvar = tk.BooleanVar()
excel_check = tk.Checkbutton(save_frame, variable=excelvar, onvalue=1,
                             offvalue=0)
excel_check.grid(row=2, column=0, padx=10, pady=10, sticky='E')
excel_lbl = tk.Label(save_frame, text=".xlsx", font=(aaa, 13))
excel_lbl.grid(row=2, column=1, padx=10, pady=10, sticky='W')

jsonvar = tk.BooleanVar()
json_check = tk.Checkbutton(save_frame, variable=jsonvar, onvalue=1,
                            offvalue=0)
json_check.grid(row=3, column=0, padx=10, pady=10, sticky='E')
json_lbl = tk.Label(save_frame, text=".json", font=(aaa, 13))
json_lbl.grid(row=3, column=1, padx=10, pady=10, sticky='W')

txtvar = tk.BooleanVar()
txt_check = tk.Checkbutton(save_frame, variable=txtvar, onvalue=1,
                           offvalue=0)
txt_check.grid(row=4, column=0, padx=10, pady=10, sticky='E')
txt_lbl = tk.Label(save_frame, text=".txt", font=(aaa, 13))
txt_lbl.grid(row=4, column=1, padx=10, pady=10, sticky='W')

# final submit button
scraping_btn = tk.Button(window, text="SCRAPE IT!", command=scrape_master,
                         bg='#fcdb03', font=(aaa, 20))
scraping_btn.grid(row=5, column=0, columnspan=3, padx=10, pady=(15, 25))

# CLOSING WINDOW LOOP
window.mainloop()
